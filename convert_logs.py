#!/usr/bin/env python3
"""从日志填报 Excel 提取指定年月的日志，生成统计表格式。"""

from __future__ import annotations

import argparse
import calendar
import io
import json
import re
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUTPUT_COLUMNS = ["项目编号", "项目", "姓名", "日期", "工时（小时）", "加班", "工作内容"]
DEFAULT_MAPPING_FILE = Path(__file__).with_name("project_mapping.json")
SKIP_PROJECTS = {"请假"}
SKIP_CONTENTS = {"请假"}
MAX_PROJECTS = 10
UNKNOWN_PROJECT = "未知项目"


def load_project_mapping(mapping_file: Path | None) -> dict[str, str]:
    from project_store import load_flat_mapping

    return load_flat_mapping(mapping_file)


def resolve_project(project: str | None, mapping: dict[str, str]) -> tuple[str, str | None]:
    """返回 (标准项目名, 原始项目名)。未找到映射时归入未知项目。"""
    if not project or str(project).strip() == "":
        return UNKNOWN_PROJECT, None

    original = str(project).strip()
    if original in SKIP_PROJECTS:
        return original, original

    if original in mapping:
        return mapping[original], original

    lowered = original.lower()
    for key, value in mapping.items():
        if key.lower() == lowered:
            return value, original

    return UNKNOWN_PROJECT, original


def parse_sheet_date(sheet_name: str, year: int, month: int) -> datetime | None:
    match = re.fullmatch(r"(\d{1,2})\.(\d{1,2})", sheet_name.strip())
    if not match:
        return None
    sheet_month, day = int(match.group(1)), int(match.group(2))
    if sheet_month != month:
        return None
    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def parse_sheet_date_for_year(sheet_name: str, year: int) -> datetime | None:
    match = re.fullmatch(r"(\d{1,2})\.(\d{1,2})", sheet_name.strip())
    if not match:
        return None
    month, day = int(match.group(1)), int(match.group(2))
    if not 1 <= month <= 12:
        return None
    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    return str(value).strip() == ""


def parse_hours(value) -> float | None:
    if is_blank(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def detect_project_slots(df: pd.DataFrame) -> int:
    if len(df) == 0:
        return MAX_PROJECTS

    header = [str(v).strip() if not is_blank(v) else "" for v in df.iloc[0].tolist()]
    slot_count = sum(1 for cell in header if cell.startswith("项目"))
    if slot_count > 0:
        return min(slot_count, MAX_PROJECTS)

    col_count = df.shape[1]
    return min(max((col_count - 1) // 3, 1), MAX_PROJECTS)


def extract_entries_from_sheet(df: pd.DataFrame, log_date: datetime, mapping: dict[str, str]) -> list[dict]:
    records: list[dict] = []
    project_slots = detect_project_slots(df)

    for row_index in range(1, len(df)):
        row = df.iloc[row_index].tolist()
        required_len = 1 + project_slots * 3
        while len(row) < required_len:
            row.append(None)

        name = row[0]
        if is_blank(name):
            continue
        name = str(name).strip()

        for slot in range(project_slots):
            project = row[1 + slot * 3]
            content = row[2 + slot * 3]
            hours = row[3 + slot * 3]

            hours_value = parse_hours(hours)
            if hours_value is None:
                continue
            if is_blank(project) and is_blank(content):
                continue

            project_text = "" if is_blank(project) else str(project).strip()
            content_text = "" if is_blank(content) else str(content).strip()

            if project_text in SKIP_PROJECTS or content_text in SKIP_CONTENTS:
                continue

            resolved_project, original_project = resolve_project(project_text, mapping)
            if resolved_project == UNKNOWN_PROJECT and original_project:
                content_text = f"【原项目: {original_project}】{content_text}"

            records.append(
                {
                    "项目编号": None,
                    "项目": resolved_project,
                    "姓名": name,
                    "日期": log_date,
                    "工时（小时）": hours_value,
                    "加班": None,
                    "工作内容": content_text,
                }
            )
    return records


def collect_logs_for_year(source_file: Path | str | io.BytesIO, year: int, mapping: dict[str, str]) -> list[dict]:
    excel = pd.ExcelFile(source_file, engine="calamine")
    all_records: list[dict] = []

    for sheet_name in excel.sheet_names:
        if sheet_name == "目录":
            continue
        log_date = parse_sheet_date_for_year(sheet_name, year)
        if log_date is None:
            continue

        df = pd.read_excel(source_file, sheet_name=sheet_name, header=None, engine="calamine")
        all_records.extend(extract_entries_from_sheet(df, log_date, mapping))

    all_records.sort(
        key=lambda item: (
            item["日期"],
            1 if item["项目"] == UNKNOWN_PROJECT else 0,
            item["项目"],
            item["姓名"],
            item["工作内容"],
        )
    )
    return all_records


def collect_logs(source_file: Path | str | io.BytesIO, year: int, month: int, mapping: dict[str, str]) -> list[dict]:
    excel = pd.ExcelFile(source_file, engine="calamine")
    all_records: list[dict] = []

    for sheet_name in excel.sheet_names:
        if sheet_name == "目录":
            continue
        log_date = parse_sheet_date(sheet_name, year, month)
        if log_date is None:
            continue

        df = pd.read_excel(source_file, sheet_name=sheet_name, header=None, engine="calamine")
        all_records.extend(extract_entries_from_sheet(df, log_date, mapping))

    all_records.sort(
        key=lambda item: (
            1 if item["项目"] == UNKNOWN_PROJECT else 0,
            item["项目"],
            item["姓名"],
            item["日期"],
            item["工作内容"],
        )
    )
    return all_records


def serialize_record(record: dict) -> dict[str, Any]:
    date_value = record["日期"]
    return {
        "项目编号": record["项目编号"],
        "项目": record["项目"],
        "姓名": record["姓名"],
        "日期": date_value.strftime("%Y-%m-%d") if isinstance(date_value, datetime) else str(date_value),
        "工时（小时）": record["工时（小时）"],
        "加班": record["加班"],
        "工作内容": record["工作内容"],
    }


def compute_summary(records: list[dict]) -> dict[str, Any]:
    if not records:
        return {
            "total_records": 0,
            "total_hours": 0,
            "people_count": 0,
            "days_count": 0,
            "project_count": 0,
            "unknown_count": 0,
            "unknown_hours": 0,
            "avg_hours_per_person": 0,
            "by_project": [],
            "by_person": [],
        }

    total_hours = sum(float(r["工时（小时）"]) for r in records)
    people = sorted({r["姓名"] for r in records})
    days = sorted({r["日期"].date() if isinstance(r["日期"], datetime) else r["日期"] for r in records})
    projects = sorted({r["项目"] for r in records})

    project_hours: dict[str, float] = {}
    person_hours: dict[str, float] = {}
    for record in records:
        hours = float(record["工时（小时）"])
        project_hours[record["项目"]] = project_hours.get(record["项目"], 0) + hours
        person_hours[record["姓名"]] = person_hours.get(record["姓名"], 0) + hours

    unknown_records = [r for r in records if r["项目"] == UNKNOWN_PROJECT]
    unknown_hours = sum(float(r["工时（小时）"]) for r in unknown_records)

    by_project = [
        {"项目": name, "工时": round(hours, 2), "占比": round(hours / total_hours * 100, 1) if total_hours else 0}
        for name, hours in sorted(project_hours.items(), key=lambda item: item[1], reverse=True)
    ]
    by_person = [
        {"姓名": name, "工时": round(hours, 2), "占比": round(hours / total_hours * 100, 1) if total_hours else 0}
        for name, hours in sorted(person_hours.items(), key=lambda item: item[1], reverse=True)
    ]

    return {
        "total_records": len(records),
        "total_hours": round(total_hours, 2),
        "people_count": len(people),
        "days_count": len(days),
        "project_count": len(projects),
        "unknown_count": len(unknown_records),
        "unknown_hours": round(unknown_hours, 2),
        "avg_hours_per_person": round(total_hours / len(people), 2) if people else 0,
        "by_project": by_project,
        "by_person": by_person,
    }


def build_convert_result(
    source_file: Path | str | io.BytesIO,
    year: int,
    month: int,
    mapping_file: Path | None = None,
    mapping: dict[str, str] | None = None,
) -> dict[str, Any]:
    project_mapping = mapping if mapping is not None else load_project_mapping(mapping_file)
    records = collect_logs(source_file, year, month, project_mapping)
    if not records:
        raise ValueError(f"{year} 年 {month} 月未找到任何日志记录")

    token = uuid.uuid4().hex
    sheet_name = build_default_sheet_name(year, month)
    return {
        "token": token,
        "year": year,
        "month": month,
        "sheet_name": sheet_name,
        "filename": build_default_output_name(year, month),
        "records": records,
        "rows": [serialize_record(record) for record in records],
        "summary": compute_summary(records),
    }


def autosize_columns(ws) -> None:
    for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        max_length = len(column_name)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is None:
                continue
            if isinstance(value, datetime):
                text = value.strftime("%Y-%m-%d")
            else:
                text = str(value)
            max_length = max(max_length, min(len(text), 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2


def write_output(records: list[dict], output_file: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True)
    for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, record in enumerate(records, start=2):
        for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
            value = record[column_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if column_name == "日期" and isinstance(value, datetime):
                cell.number_format = "YYYY-MM-DD"
            if column_name == "工作内容":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)


def records_to_excel_bytes(records: list[dict], sheet_name: str) -> bytes:
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True)
    for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, record in enumerate(records, start=2):
        for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
            value = record[column_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if column_name == "日期" and isinstance(value, datetime):
                cell.number_format = "YYYY-MM-DD"
            if column_name == "工作内容":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    wb.save(buffer)
    return buffer.getvalue()


def build_default_output_name(year: int, month: int) -> str:
    return f"{month}月日志统计表1-31.xlsx"


def build_default_sheet_name(year: int, month: int) -> str:
    last_day = calendar.monthrange(year, month)[1]
    return f"{month}.1-{month}.{last_day}"


def prompt_year_month() -> tuple[int, int]:
    while True:
        year_text = input("请输入年份（例如 2026）: ").strip()
        month_text = input("请输入月份（例如 5）: ").strip()
        try:
            year = int(year_text)
            month = int(month_text)
        except ValueError:
            print("输入无效，请重新输入。")
            continue
        if not 1 <= month <= 12:
            print("月份必须在 1-12 之间。")
            continue
        return year, month


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从日志填报 Excel 生成统计表格式 Excel")
    parser.add_argument("year", type=int, nargs="?", help="年份，例如 2026")
    parser.add_argument("month", type=int, nargs="?", help="月份，例如 5")
    parser.add_argument(
        "-i",
        "--input",
        dest="input_file",
        type=Path,
        default=Path("4.27-7.30日志填报.xlsx"),
        help="日志填报 Excel 路径",
    )
    parser.add_argument(
        "-o",
        "--output",
        dest="output_file",
        type=Path,
        help="输出 Excel 路径，默认自动生成",
    )
    parser.add_argument(
        "-m",
        "--mapping",
        dest="mapping_file",
        type=Path,
        default=DEFAULT_MAPPING_FILE,
        help="项目名称映射 JSON 文件",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.year is None or args.month is None:
        year, month = prompt_year_month()
    else:
        year, month = args.year, args.month

    if not 1 <= month <= 12:
        print("错误: 月份必须在 1-12 之间", file=sys.stderr)
        return 1
    if not args.input_file.exists():
        print(f"错误: 找不到输入文件 {args.input_file}", file=sys.stderr)
        return 1

    try:
        result = build_convert_result(args.input_file, year, month, args.mapping_file)
    except ValueError as exc:
        print(f"警告: {exc}", file=sys.stderr)
        return 1

    output_file = args.output_file or Path(result["filename"])
    write_output(result["records"], output_file, result["sheet_name"])

    summary = result["summary"]
    print(f"已生成: {output_file}")
    print(
        f"共 {summary['total_records']} 条记录，"
        f"{summary['people_count']} 人，"
        f"{summary['days_count']} 天"
    )
    if summary["unknown_count"]:
        print(f"提示: 有 {summary['unknown_count']} 条记录归入「{UNKNOWN_PROJECT}」，可在 project_mapping.json 中补充映射。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
